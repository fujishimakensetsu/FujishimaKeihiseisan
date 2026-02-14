"""
エクスポートルーター
CSV/Excel/PDF出力機能
選択エクスポート対応
"""
import os
import time
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from jose import JWTError, jwt
from database import db
from services.auth_service import get_current_user_optional, get_current_user
import config

router = APIRouter()

def wrap_text_every_n(text: str, n: int = 7) -> str:
    """n文字ごとに改行を挿入する"""
    if not text or len(text) <= n:
        return text
    return "\n".join(text[i:i+n] for i in range(0, len(text), n))

def format_date_slash(date_str: str) -> str:
    """日付をスラッシュ区切りに変換 (YYYY-MM-DD → YYYY/M/D)"""
    if not date_str:
        return ""
    try:
        parts = date_str.split("-")
        if len(parts) == 3:
            year = int(parts[0])
            month = int(parts[1])
            day = int(parts[2])
            return f"{year}/{month}/{day}"
    except (ValueError, IndexError):
        pass
    return date_str

# 日本語フォント管理
def download_japanese_font():
    """Noto Sans JPフォントをダウンロード"""
    font_path = os.path.join(config.FONT_DIR, "NotoSansJP-Regular.ttf")

    if os.path.exists(font_path):
        print("日本語フォントは既にダウンロード済みです")
        return font_path

    print("日本語フォントをダウンロード中...")
    try:
        import requests
        url = "https://github.com/google/fonts/raw/main/ofl/notosansjp/NotoSansJP%5Bwght%5D.ttf"
        response = requests.get(url, timeout=30)

        if response.status_code == 200:
            with open(font_path, "wb") as f:
                f.write(response.content)
            print("日本語フォントのダウンロードが完了しました")
            return font_path
        else:
            print(f"フォントのダウンロードに失敗しました: {response.status_code}")
            return None
    except Exception as e:
        print(f"フォントダウンロードエラー: {e}")
        return None

# フォントを事前にダウンロード
JAPANESE_FONT_PATH = download_japanese_font()

@router.get("/api/export/csv")
async def export_csv(token: Optional[str] = None, u_id: Optional[str] = Depends(get_current_user_optional)):
    """CSV出力（サブコレクション対応）"""
    import pandas as pd

    # トークンパラメータがある場合はそれを使用
    if token:
        try:
            payload = jwt.decode(token, config.SECRET_KEY, algorithms=[config.ALGORITHM])
            u_id = payload.get("sub")
        except JWTError:
            raise HTTPException(status_code=401, detail="無効なトークンです")

    if not u_id:
        raise HTTPException(status_code=401, detail="認証が必要です")

    # サブコレクションからレコードを取得
    records_ref = db.collection(config.COL_USERS).document(u_id).collection("records").stream()
    records = [r.to_dict() for r in records_ref]

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    df = pd.DataFrame(records)
    csv_path = f"{config.UPLOAD_DIR}/export_{u_id}_{int(time.time())}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    return FileResponse(csv_path, media_type="text/csv", filename=f"receipts_{u_id}.csv")

@router.get("/api/export/excel")
async def export_excel(token: Optional[str] = None, u_id: Optional[str] = Depends(get_current_user_optional)):
    """Excel出力（テンプレート使用・店舗名集計・駐車場合算・交通費別欄）"""
    import openpyxl
    from openpyxl.styles import Alignment
    from collections import defaultdict
    from datetime import datetime

    # トークンパラメータがある場合はそれを使用
    if token:
        try:
            payload = jwt.decode(token, config.SECRET_KEY, algorithms=[config.ALGORITHM])
            u_id = payload.get("sub")
        except JWTError:
            raise HTTPException(status_code=401, detail="無効なトークンです")

    if not u_id:
        raise HTTPException(status_code=401, detail="認証が必要です")

    # サブコレクションからレコードを取得
    records_ref = db.collection(config.COL_USERS).document(u_id).collection("records").stream()
    records = [r.to_dict() for r in records_ref]

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    # 駐輪場キーワード（駐車場から除外）
    bicycle_keywords = ["駐輪", "自転車", "サイクル", "bicycle", "cycle"]
    # 駐車場キーワード（フォールバック用）
    parking_keywords = ["駐車", "パーキング", "コインパ", "parking", "P代"]

    def is_bicycle_parking(record):
        """駐輪場かどうかを判定"""
        vendor = record.get("vendor_name", "").lower()
        for kw in bicycle_keywords:
            if kw.lower() in vendor:
                return True
        return False

    def is_parking(record):
        """駐車場かどうかを判定（Gemini解析結果のフラグを優先、駐輪場は除外）"""
        # 駐輪場は駐車場に含めない
        if is_bicycle_parking(record):
            return False
        # Geminiのis_parkingフラグを優先
        if record.get("is_parking") == True:
            return True
        # フォールバック: キーワードマッチング
        vendor = record.get("vendor_name", "").lower()
        for kw in parking_keywords:
            if kw.lower() in vendor:
                return True
        return False

    def is_ic_transport(record):
        """ICカード交通費かどうかを判定

        Gemini解析で「ICカード交通費」という文言が書類に明示的に
        記載されている場合のみtrueとなる。
        通常のレシートや交通系ICカードの利用履歴はfalse。
        """
        return record.get("is_ic_transport") == True

    # レコードを分類
    parking_records = []
    transport_records = []
    other_records = []

    for record in records:
        if is_parking(record):
            parking_records.append(record)
        elif is_ic_transport(record):
            transport_records.append(record)
        else:
            other_records.append(record)

    # 駐車場代を全合算
    parking_total = sum(r.get("total_amount", 0) for r in parking_records)
    parking_dates = sorted([r.get("date", "") for r in parking_records if r.get("date")], reverse=True)

    # その他のレコードを店舗名で集計
    vendor_totals = defaultdict(lambda: {"amount": 0, "dates": [], "category": ""})
    for record in other_records:
        vendor_name = record.get("vendor_name", "不明")
        amount = record.get("total_amount", 0)
        date = record.get("date", "")
        category = record.get("category", "その他")

        vendor_totals[vendor_name]["amount"] += amount
        if date and date not in vendor_totals[vendor_name]["dates"]:
            vendor_totals[vendor_name]["dates"].append(date)
        if not vendor_totals[vendor_name]["category"]:
            vendor_totals[vendor_name]["category"] = category

    # テンプレートファイルを読み込み
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "template.xlsx")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="テンプレートファイルが見つかりません")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # B2: 年月日を6桁で入力（YYMMDD）
    ws.cell(row=2, column=2, value=datetime.now().strftime("%y%m%d"))

    # === 経費精算欄にデータを書き込み ===

    # 駐車場代は10行目のS列のみに出力（合算金額のみ）
    if parking_total > 0:
        ws.cell(row=10, column=19, value=parking_total)  # S10: 駐車場代合算

    # その他のレコードは11行目から出力（店舗名集計済み）
    row = 11
    for vendor_name, data in sorted(vendor_totals.items(), key=lambda x: x[1]["amount"], reverse=True):
        if row > 29:
            break

        dates = sorted(data["dates"], reverse=True)
        date_str = format_date_slash(dates[0]) if dates else ""

        ws.cell(row=row, column=2, value=date_str)  # B列: 支払日
        cell_e = ws.cell(row=row, column=5, value=wrap_text_every_n(vendor_name, 7))  # E列: 支払先
        cell_e.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=8, value=data["category"])  # H列: 支払事由
        ws.cell(row=row, column=19, value=data["amount"])  # S列: 支払額（10%）
        row += 1

    # === ICカード交通費欄（Z-AE列）にデータを書き込み ===
    # ※「ICカード交通費」という文言が明示的に記載されたPDFのデータのみ出力
    # ※通常のレシートや交通系ICカードの利用履歴はここには出力されない
    # ※レコード内にitemsがある場合は、各項目を個別の行として出力する
    transport_row = 7  # 7行目から開始

    # transport_recordsを展開して、各項目を個別に出力
    transport_items = []
    for record in transport_records:
        items = record.get("items", [])
        if items and isinstance(items, list):
            # itemsがある場合は各項目を展開
            for item in items:
                transport_items.append({
                    "date": item.get("date", record.get("date", "")),
                    "vendor": item.get("vendor_name", item.get("vendor", record.get("vendor_name", ""))),
                    "amount": item.get("total_amount", item.get("amount", 0)),
                    "from_station": item.get("from_station", item.get("from", "")),
                    "to_station": item.get("to_station", item.get("to", ""))
                })
        else:
            # itemsがない場合はレコード自体を1項目として出力
            transport_items.append({
                "date": record.get("date", ""),
                "vendor": record.get("vendor_name", ""),
                "amount": record.get("total_amount", 0),
                "from_station": record.get("from_station", ""),
                "to_station": record.get("to_station", "")
            })

    # 日付順にソートして出力
    for item in sorted(transport_items, key=lambda x: x.get("date", "")):
        if transport_row > 24:  # 最大18件（7行目〜24行目）
            break

        ws.cell(row=transport_row, column=26, value=item.get("date", ""))  # Z列: 利用日
        ws.cell(row=transport_row, column=27, value=item.get("vendor", ""))  # AA列: 利用先
        ws.cell(row=transport_row, column=28, value=item.get("from_station", ""))  # AB列: 区間始まり
        ws.cell(row=transport_row, column=30, value=item.get("to_station", ""))  # AD列: 区間終わり
        ws.cell(row=transport_row, column=31, value=item.get("amount", 0))  # AE列: 利用金額

        transport_row += 1

    # 提出日を設定（G31）
    today = datetime.now().strftime("%Y/%m/%d")
    ws.cell(row=31, column=7, value=today)

    # 出力ファイルを保存
    excel_path = f"{config.UPLOAD_DIR}/export_{u_id}_{int(time.time())}.xlsx"
    wb.save(excel_path)
    wb.close()

    # ファイル名を出力日のYYMMDD形式にする（例: 260209）
    filename_date = datetime.now().strftime("%y%m%d")
    return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"{filename_date}.xlsx")

@router.get("/api/export/pdf")
async def export_pdf(token: Optional[str] = None, u_id: Optional[str] = Depends(get_current_user_optional)):
    """PDF出力（サブコレクション対応）"""
    from fpdf import FPDF

    # トークンパラメータがある場合はそれを使用
    if token:
        try:
            payload = jwt.decode(token, config.SECRET_KEY, algorithms=[config.ALGORITHM])
            u_id = payload.get("sub")
        except JWTError:
            raise HTTPException(status_code=401, detail="無効なトークンです")

    if not u_id:
        raise HTTPException(status_code=401, detail="認証が必要です")

    # サブコレクションからレコードを取得
    records_ref = db.collection(config.COL_USERS).document(u_id).collection("records").stream()
    records = [r.to_dict() for r in records_ref]

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    # 日付でソート
    records.sort(key=lambda x: x.get("date", ""), reverse=True)

    pdf = FPDF()
    pdf.add_page()

    # 日本語フォントを追加
    if JAPANESE_FONT_PATH and os.path.exists(JAPANESE_FONT_PATH):
        pdf.add_font("NotoSansJP", "", JAPANESE_FONT_PATH, uni=True)
        pdf.set_font("NotoSansJP", size=10)
    else:
        pdf.set_font("Arial", size=10)

    # タイトル
    pdf.set_font_size(16)
    pdf.cell(0, 10, "領収書一覧", ln=True, align="C")
    pdf.ln(5)

    # ヘッダー
    pdf.set_font_size(10)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(30, 8, "日付", border=1, fill=True)
    pdf.cell(80, 8, "店舗名", border=1, fill=True)
    pdf.cell(40, 8, "金額", border=1, fill=True, align="R")
    pdf.cell(40, 8, "カテゴリ", border=1, fill=True)
    pdf.ln()

    # データ行
    for record in records:
        date = record.get("date", "")
        vendor = record.get("vendor_name", "")[:25]
        amount = f"¥{record.get('total_amount', 0):,}"
        category = record.get("category", "その他")

        # 交互に背景色を変更
        if records.index(record) % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
            fill = True
        else:
            fill = False

        pdf.cell(30, 8, date, border=1, fill=fill)
        pdf.cell(80, 8, vendor, border=1, fill=fill)
        pdf.cell(40, 8, amount, border=1, fill=fill, align="R")
        pdf.cell(40, 8, category, border=1, fill=fill)
        pdf.ln()

    # 合計金額を計算
    total = sum([record.get("total_amount", 0) for record in records])
    pdf.ln(5)
    pdf.set_font_size(12)
    pdf.cell(110, 10, "合計金額:", align="R")
    pdf.set_font_size(14)
    pdf.cell(40, 10, f"¥{total:,}", align="R")

    pdf_path = f"{config.UPLOAD_DIR}/export_{u_id}_{int(time.time())}.pdf"
    pdf.output(pdf_path)

    return FileResponse(pdf_path, media_type="application/pdf", filename=f"receipts_{u_id}.pdf")

# ========== 選択エクスポート機能 ==========

@router.post("/api/export/selected/csv")
async def export_selected_csv(data: dict, u_id: str = Depends(get_current_user)):
    """選択したレコードのみCSV出力"""
    import pandas as pd

    record_ids = data.get("record_ids", [])

    if not record_ids:
        raise HTTPException(status_code=400, detail="エクスポートするレコードが指定されていません")

    print(f"=== Selected CSV export ===")
    print(f"User: {u_id}, Records: {len(record_ids)} items")

    # 選択したレコードを取得
    records = []
    for record_id in record_ids:
        doc_ref = db.collection(config.COL_USERS).document(u_id).collection("records").document(record_id)
        doc = doc_ref.get()
        if doc.exists:
            records.append(doc.to_dict())

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    df = pd.DataFrame(records)
    csv_path = f"{config.UPLOAD_DIR}/export_selected_{u_id}_{int(time.time())}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    return FileResponse(csv_path, media_type="text/csv", filename=f"receipts_selected_{u_id}.csv")

@router.post("/api/export/selected/excel")
async def export_selected_excel(data: dict, u_id: str = Depends(get_current_user)):
    """選択したレコードのみExcel出力（テンプレート使用・駐車場合算・交通費別欄）"""
    import openpyxl
    from openpyxl.styles import Alignment
    from collections import defaultdict
    from datetime import datetime

    record_ids = data.get("record_ids", [])

    if not record_ids:
        raise HTTPException(status_code=400, detail="エクスポートするレコードが指定されていません")

    print(f"=== Selected Excel export ===")
    print(f"User: {u_id}, Records: {len(record_ids)} items")

    # 選択したレコードを取得
    records = []
    for record_id in record_ids:
        doc_ref = db.collection(config.COL_USERS).document(u_id).collection("records").document(record_id)
        doc = doc_ref.get()
        if doc.exists:
            records.append(doc.to_dict())

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    # 駐輪場キーワード（駐車場から除外）
    bicycle_keywords = ["駐輪", "自転車", "サイクル", "bicycle", "cycle"]
    # 駐車場キーワード（フォールバック用）
    parking_keywords = ["駐車", "パーキング", "コインパ", "parking", "P代"]

    def is_bicycle_parking(record):
        """駐輪場かどうかを判定"""
        vendor = record.get("vendor_name", "").lower()
        for kw in bicycle_keywords:
            if kw.lower() in vendor:
                return True
        return False

    def is_parking(record):
        """駐車場かどうかを判定（Gemini解析結果のフラグを優先、駐輪場は除外）"""
        # 駐輪場は駐車場に含めない
        if is_bicycle_parking(record):
            return False
        # Geminiのis_parkingフラグを優先
        if record.get("is_parking") == True:
            return True
        # フォールバック: キーワードマッチング
        vendor = record.get("vendor_name", "").lower()
        for kw in parking_keywords:
            if kw.lower() in vendor:
                return True
        return False

    def is_ic_transport(record):
        """ICカード交通費かどうかを判定

        Gemini解析で「ICカード交通費」という文言が書類に明示的に
        記載されている場合のみtrueとなる。
        通常のレシートや交通系ICカードの利用履歴はfalse。
        """
        return record.get("is_ic_transport") == True

    # レコードを分類
    parking_records = []
    transport_records = []
    other_records = []

    for record in records:
        if is_parking(record):
            parking_records.append(record)
        elif is_ic_transport(record):
            transport_records.append(record)
        else:
            other_records.append(record)

    # 駐車場代を全合算
    parking_total = sum(r.get("total_amount", 0) for r in parking_records)
    parking_dates = sorted([r.get("date", "") for r in parking_records if r.get("date")], reverse=True)

    # その他のレコードを店舗名で集計
    vendor_totals = defaultdict(lambda: {"amount": 0, "dates": [], "category": ""})
    for record in other_records:
        vendor_name = record.get("vendor_name", "不明")
        amount = record.get("total_amount", 0)
        date = record.get("date", "")
        category = record.get("category", "その他")

        vendor_totals[vendor_name]["amount"] += amount
        if date and date not in vendor_totals[vendor_name]["dates"]:
            vendor_totals[vendor_name]["dates"].append(date)
        if not vendor_totals[vendor_name]["category"]:
            vendor_totals[vendor_name]["category"] = category

    # テンプレートファイルを読み込み
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "template.xlsx")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="テンプレートファイルが見つかりません")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # B2: 年月日を6桁で入力（YYMMDD）
    ws.cell(row=2, column=2, value=datetime.now().strftime("%y%m%d"))

    # === 経費精算欄にデータを書き込み ===

    # 駐車場代は10行目のS列のみに出力（合算金額のみ）
    if parking_total > 0:
        ws.cell(row=10, column=19, value=parking_total)  # S10: 駐車場代合算

    # その他のレコードは11行目から出力（店舗名集計済み）
    row = 11
    for vendor_name, data in sorted(vendor_totals.items(), key=lambda x: x[1]["amount"], reverse=True):
        if row > 29:
            break

        dates = sorted(data["dates"], reverse=True)
        date_str = format_date_slash(dates[0]) if dates else ""

        ws.cell(row=row, column=2, value=date_str)  # B列: 支払日
        cell_e = ws.cell(row=row, column=5, value=wrap_text_every_n(vendor_name, 7))  # E列: 支払先
        cell_e.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=8, value=data["category"])  # H列: 支払事由
        ws.cell(row=row, column=19, value=data["amount"])  # S列: 支払額（10%）
        row += 1

    # === ICカード交通費欄（Z-AE列）にデータを書き込み ===
    # ※レコード内にitemsがある場合は、各項目を個別の行として出力する
    transport_row = 7  # 7行目から開始

    # transport_recordsを展開して、各項目を個別に出力
    transport_items = []
    for record in transport_records:
        items = record.get("items", [])
        if items and isinstance(items, list):
            # itemsがある場合は各項目を展開
            for item in items:
                transport_items.append({
                    "date": item.get("date", record.get("date", "")),
                    "vendor": item.get("vendor_name", item.get("vendor", record.get("vendor_name", ""))),
                    "amount": item.get("total_amount", item.get("amount", 0)),
                    "from_station": item.get("from_station", item.get("from", "")),
                    "to_station": item.get("to_station", item.get("to", ""))
                })
        else:
            # itemsがない場合はレコード自体を1項目として出力
            transport_items.append({
                "date": record.get("date", ""),
                "vendor": record.get("vendor_name", ""),
                "amount": record.get("total_amount", 0),
                "from_station": record.get("from_station", ""),
                "to_station": record.get("to_station", "")
            })

    # 日付順にソートして出力（降順）
    for item in sorted(transport_items, key=lambda x: x.get("date", ""), reverse=True):
        if transport_row > 24:  # 最大18件（7行目〜24行目）
            break

        ws.cell(row=transport_row, column=26, value=item.get("date", ""))  # Z列: 利用日
        ws.cell(row=transport_row, column=27, value=item.get("vendor", ""))  # AA列: 利用先
        ws.cell(row=transport_row, column=28, value=item.get("from_station", ""))  # AB列: 区間始まり
        ws.cell(row=transport_row, column=30, value=item.get("to_station", ""))  # AD列: 区間終わり
        ws.cell(row=transport_row, column=31, value=item.get("amount", 0))  # AE列: 利用金額

        transport_row += 1

    # 提出日を設定（G31）
    today = datetime.now().strftime("%Y/%m/%d")
    ws.cell(row=31, column=7, value=today)

    # 出力ファイルを保存
    excel_path = f"{config.UPLOAD_DIR}/export_selected_{u_id}_{int(time.time())}.xlsx"
    wb.save(excel_path)
    wb.close()

    # ファイル名を出力日のYYMMDD形式にする（例: 260209）
    filename_date = datetime.now().strftime("%y%m%d")
    return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"{filename_date}.xlsx")

@router.post("/api/export/selected/pdf")
async def export_selected_pdf(data: dict, u_id: str = Depends(get_current_user)):
    """選択したレコードのみPDF出力"""
    from fpdf import FPDF

    record_ids = data.get("record_ids", [])

    if not record_ids:
        raise HTTPException(status_code=400, detail="エクスポートするレコードが指定されていません")

    print(f"=== Selected PDF export ===")
    print(f"User: {u_id}, Records: {len(record_ids)} items")

    # 選択したレコードを取得
    records = []
    for record_id in record_ids:
        doc_ref = db.collection(config.COL_USERS).document(u_id).collection("records").document(record_id)
        doc = doc_ref.get()
        if doc.exists:
            records.append(doc.to_dict())

    if not records:
        raise HTTPException(status_code=404, detail="データがありません")

    # 日付でソート
    records.sort(key=lambda x: x.get("date", ""), reverse=True)

    pdf = FPDF()
    pdf.add_page()

    # 日本語フォントを追加
    if JAPANESE_FONT_PATH and os.path.exists(JAPANESE_FONT_PATH):
        pdf.add_font("NotoSansJP", "", JAPANESE_FONT_PATH, uni=True)
        pdf.set_font("NotoSansJP", size=10)
    else:
        pdf.set_font("Arial", size=10)

    # タイトル
    pdf.set_font_size(16)
    pdf.cell(0, 10, "領収書一覧（選択分）", ln=True, align="C")
    pdf.ln(5)

    # ヘッダー
    pdf.set_font_size(10)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(30, 8, "日付", border=1, fill=True)
    pdf.cell(80, 8, "店舗名", border=1, fill=True)
    pdf.cell(40, 8, "金額", border=1, fill=True, align="R")
    pdf.cell(40, 8, "カテゴリ", border=1, fill=True)
    pdf.ln()

    # データ行
    for idx, record in enumerate(records):
        date = record.get("date", "")
        vendor = record.get("vendor_name", "")[:25]
        amount = f"¥{record.get('total_amount', 0):,}"
        category = record.get("category", "その他")

        if idx % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
            fill = True
        else:
            fill = False

        pdf.cell(30, 8, date, border=1, fill=fill)
        pdf.cell(80, 8, vendor, border=1, fill=fill)
        pdf.cell(40, 8, amount, border=1, fill=fill, align="R")
        pdf.cell(40, 8, category, border=1, fill=fill)
        pdf.ln()

    # 合計金額を計算
    total = sum([record.get("total_amount", 0) for record in records])
    pdf.ln(5)
    pdf.set_font_size(12)
    pdf.cell(110, 10, "合計金額:", align="R")
    pdf.set_font_size(14)
    pdf.cell(40, 10, f"¥{total:,}", align="R")

    pdf_path = f"{config.UPLOAD_DIR}/export_selected_{u_id}_{int(time.time())}.pdf"
    pdf.output(pdf_path)

    return FileResponse(pdf_path, media_type="application/pdf", filename=f"receipts_selected_{u_id}.pdf")
